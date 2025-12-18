import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
import io, zipfile, tempfile, os

st.set_page_config(page_title="Master Excel Utility Tool", layout="wide")

# =====================================================
# SIDEBAR
# =====================================================
st.sidebar.title("🛠 Master Utility Tool")
menu = st.sidebar.radio(
    "Select Tool",
    [
        "📂 New File Creation",
        "🔄 CSV → XLSX Converter",
        "📝 Merchant Auto Rename",
        "📊 Excel Formula Updater",
        "🚚 Courier Cost Updater"
    ]
)

st.sidebar.info(
    "☁ Cloud Version\n"
    "📤 Upload Files / Folder (ZIP)\n"
    "📥 Download Output (ZIP)\n\n"
    "❌ No path paste required"
)

# =====================================================
# HELPERS
# =====================================================
def extract_files(files, zip_file, exts):
    extracted = []

    if files:
        extracted.extend(files)

    if zip_file:
        tmp = tempfile.mkdtemp()
        zip_path = os.path.join(tmp, zip_file.name)
        with open(zip_path, "wb") as f:
            f.write(zip_file.read())

        with zipfile.ZipFile(zip_path) as z:
            z.extractall(tmp)

        for root, _, filenames in os.walk(tmp):
            for name in filenames:
                if name.lower().endswith(exts):
                    extracted.append(open(os.path.join(root, name), "rb"))
    return extracted


def make_zip(file_buffers):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in file_buffers:
            z.writestr(name, data.getvalue())
    buf.seek(0)
    return buf


# =====================================================
# 1️⃣ NEW FILE CREATION
# =====================================================
if menu == "📂 New File Creation":
    st.title("📂 New File Creation Tool")

    adf_file = st.file_uploader("📘 Upload ADF Foods File", type="xlsx")
    pricing_file = st.file_uploader("📗 Upload Pricing Format File", type="xlsx")

    files = st.file_uploader("📤 Upload Excel Files", type="xlsx", accept_multiple_files=True)
    zip_file = st.file_uploader("📦 OR Upload Folder (ZIP)", type="zip")

    if st.button("🚀 Start Processing"):
        if not adf_file or not pricing_file:
            st.error("❌ Upload reference files")
            st.stop()

        source_files = extract_files(files, zip_file, (".xlsx",))
        if not source_files:
            st.error("❌ No Excel files found")
            st.stop()

        ref_wb = load_workbook(adf_file, data_only=False)
        ref_ws = ref_wb.active
        ref_formulas = {
            c: ref_ws.cell(2, c).value
            for c in range(38, 65)
            if isinstance(ref_ws.cell(2, c).value, str)
            and ref_ws.cell(2, c).value.startswith("=")
        }

        pricing_df = pd.read_excel(pricing_file, header=None)
        pricing_data = pricing_df.values.tolist()

        extra_cols = [
            "Weight","Merchant Zone","Courier Name","FWD","CHECK","COD","CHECK",
            "RTO","CHECK","REVERSAL","CHECK","QC","TOTAL","TOTAL+GST","Diff",
            "Courier charged weight","Courier Zone","Freight","RTO","RTO Discount",
            "Reverse","COD","SDL","Fuel","QC","Others","Gross"
        ]

        output = []
        prog = st.progress(0)

        for i, f in enumerate(source_files, start=1):
            df = pd.read_excel(f, usecols=range(37))
            wb = load_workbook(f)
            ws = wb.active
            ws.title = "Raw"

            ws.delete_rows(1, ws.max_row)

            ws.append(list(df.columns))
            for _, row in df.iterrows():
                ws.append(list(row))

            for j, col in enumerate(extra_cols):
                ws.cell(1, 38 + j).value = col

            last_row = ws.max_row
            for col, formula in ref_formulas.items():
                letter = ws.cell(2, col).column_letter
                for r in range(2, last_row + 1):
                    ws[f"{letter}{r}"] = Translator(formula, f"{letter}2").translate_formula(f"{letter}{r}")

            if "Pricing" in wb.sheetnames:
                wb.remove(wb["Pricing"])
            ps = wb.create_sheet("Pricing")
            for r, row in enumerate(pricing_data, start=1):
                for c, v in enumerate(row, start=1):
                    ps.cell(r, c, v)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            output.append((f.name, buf))
            prog.progress(i / len(source_files))

        st.success("🎉 New files created")
        st.download_button("📥 Download ZIP", make_zip(output), "New_Files.zip", "application/zip")

# =====================================================
# 2️⃣ CSV → XLSX
# =====================================================
elif menu == "🔄 CSV → XLSX Converter":
    st.title("🔄 CSV → XLSX Converter")

    files = st.file_uploader("📤 Upload CSV Files", type="csv", accept_multiple_files=True)
    zip_file = st.file_uploader("📦 OR Upload CSV Folder (ZIP)", type="zip")

    if st.button("🚀 Convert"):
        csvs = extract_files(files, zip_file, (".csv",))
        if not csvs:
            st.error("❌ No CSV files")
            st.stop()

        output = []
        for f in csvs:
            df = pd.read_csv(f)
            buf = io.BytesIO()
            df.to_excel(buf, index=False)
            buf.seek(0)
            output.append((f.name.replace(".csv", ".xlsx"), buf))

        st.download_button("📥 Download XLSX ZIP", make_zip(output), "Converted_XLSX.zip", "application/zip")

# =====================================================
# 3️⃣ MERCHANT AUTO RENAME
# =====================================================
elif menu == "📝 Merchant Auto Rename":
    st.title("📝 Merchant Auto Rename")

    files = st.file_uploader("📤 Upload Excel Files", type="xlsx", accept_multiple_files=True)
    zip_file = st.file_uploader("📦 OR Upload Folder (ZIP)", type="zip")

    if st.button("🚀 Rename"):
        excels = extract_files(files, zip_file, (".xlsx",))
        output = []

        for f in excels:
            df = pd.read_excel(f, header=None)
            name = f.name
            if str(df.iloc[0,0]).lower() == "client name":
                name = str(df.iloc[1,0]).strip() + ".xlsx"

            buf = io.BytesIO()
            df.to_excel(buf, index=False, header=False)
            buf.seek(0)
            output.append((name, buf))

        st.download_button("📥 Download Renamed ZIP", make_zip(output), "Renamed_Files.zip", "application/zip")
